"""
app.py — Roteirização de Inspeção de Linhas de Transmissão
Nova arquitetura: data / domain / services / ui
JOIN obrigatório via COD_ATIVO entre VIEW_PLANO_CONSOLIDADO_INSPECAO e VW_TORRES_COM_CRITICIDADE
"""

from __future__ import annotations

import io
import time
import pandas as pd
import streamlit as st
from streamlit_folium import st_folium
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.database import (
    iniciar_device_flow, concluir_login, is_authenticated,
    tentar_login_silencioso, logout,
    load_inspecoes_consolidadas, get_filter_options,
    load_torres_por_instalacao, load_ss_por_ativos, load_ss_por_empresa,
    load_torres_com_ss_abertas, sid_atual,
)
from services.weather import get_weather, get_forecast_5d, weather_badge
from components.mapa  import build_map
from utils.routing    import (
    calcular_urgencia, clusterizar, calcular_score_hibrido,
    selecionar_os, otimizar_rota, resumo_rota,
)

# ── Enums de prioridade (usados na UI) ──────────────────────────────────────
from enum import IntEnum
import numpy as np

class Prioridade(IntEnum):
    MAXIMA = 1
    ALTA   = 2
    NORMAL = 3

# priorizar() e selecionar_inspecoes() agora vivem em utils/routing.py
# como calcular_urgencia()+calcular_score_hibrido() e selecionar_os()
def priorizar(df: pd.DataFrame) -> pd.DataFrame:
    """Wrapper de compatibilidade — executa pipeline completo de scores."""
    from utils.routing import calcular_urgencia, clusterizar, calcular_score_hibrido
    df = calcular_urgencia(df)
    df = clusterizar(df)
    df = calcular_score_hibrido(df)
    return df.sort_values(["PRIORIDADE", "SCORE"], ascending=[True, False],
                          na_position="last").reset_index(drop=True)
# ───────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────
# HELPERS DE EXPORT EXCEL
# (definidos aqui para estarem disponíveis em todo o módulo)
# ──────────────────────────────────────────────

def _estilo_base_excel():
    thin  = Side(style="thin", color="1E2330")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    return borda


def _gerar_excel_rota(df: pd.DataFrame, resumo: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rota de Inspeção"

    borda = _estilo_base_excel()

    ws.merge_cells("A1:N1")
    tc = ws["A1"]
    tc.value = f"Rota de Inspeção — {resumo.get('total_os','?')} OS | {resumo.get('distancia_total','?')} km"
    tc.font  = Font(name="Arial", bold=True, size=14, color="00CFFF")
    tc.fill  = PatternFill("solid", fgColor="0D1117")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    sc = ws["A2"]
    sc.value = f"Gerado em: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    sc.font  = Font(name="Arial", size=9, color="7A8099")
    sc.fill  = PatternFill("solid", fgColor="0D1117")
    sc.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="00CFFF")
        cell.fill      = PatternFill("solid", fgColor="1A1F2E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda
    ws.row_dimensions[3].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), 4):
        bg = "13161D" if ri % 2 == 0 else "0D0F14"
        for ci, val in enumerate(row_data, 1):
            col_name = headers[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = borda

            fg = "E8EAF0"
            bold = False
            if col_name == "Prioridade":
                if "ATRASADA" in str(val):  fg, bold = "FF6B6B", True
                elif "VENCE"  in str(val):  fg = "FFD700"
                else:                       fg = "81C784"
            elif col_name == "Atraso (d)" and str(val) not in ("0", "–", ""):
                try:
                    if int(val) > 0: fg, bold = "FF6B6B", True
                except Exception: pass

            cell.font = Font(name="Arial", size=10, color=fg, bold=bold)
        ws.row_dimensions[ri].height = 20

    for ci, h in enumerate(headers, 1):
        widths = {"Ordem":8,"OS":16,"Ativo":14,"Torre":8,"Empresa":10,"Instalação":18,
                  "Estado":16,"Status":14,"Limite":12,"Atraso (d)":12,
                  "Prioridade":18,"Score":10,"Dist→(km)":12,"Acum.(km)":12}
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)

    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gerar_excel_os(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OS Consolidadas"
    borda = _estilo_base_excel()

    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    tc = ws["A1"]
    tc.value = f"OS Consolidadas — Gerado em {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
    tc.font  = Font(name="Arial", bold=True, size=13, color="00CFFF")
    tc.fill  = PatternFill("solid", fgColor="0D1117")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="00CFFF")
        cell.fill      = PatternFill("solid", fgColor="1A1F2E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda
    ws.row_dimensions[2].height = 26

    for ri, row_data in enumerate(df.itertuples(index=False), 3):
        bg = "13161D" if ri % 2 == 0 else "0D0F14"
        for ci, val in enumerate(row_data, 1):
            col_name = headers[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = borda
            fg   = "E8EAF0"
            bold = False
            if col_name == "STATUS_PRAZO" and str(val) == "ATRASADA":
                fg, bold = "FF6B6B", True
            cell.font = Font(name="Arial", size=10, color=fg, bold=bold)
        ws.row_dimensions[ri].height = 18

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="Roteirização de Inspeção LT",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Inter:wght@300;400;600&display=swap');

    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .stApp { background: #0D0F14; color: #E8EAF0; }

    [data-testid="stSidebar"] {
        background: #13161D;
        border-right: 1px solid #1E2330;
    }
    [data-testid="metric-container"] {
        background: #13161D;
        border: 1px solid #1E2330;
        border-radius: 10px;
        padding: 16px 20px;
    }
    [data-testid="metric-container"] label {
        color: #7A8099 !important; font-size: 11px !important;
        letter-spacing: 0.08em; text-transform: uppercase;
    }
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        font-family: 'Space Mono', monospace;
        font-size: 28px !important; color: #00CFFF !important;
    }
    .app-title {
        font-family: 'Space Mono', monospace; font-size: 22px;
        font-weight: 700; color: #00CFFF; letter-spacing: -0.02em;
        border-bottom: 2px solid #1E2330; padding-bottom: 12px; margin-bottom: 20px;
    }
    .badge-atrasada { background:#FF2D2D22;color:#FF6B6B;padding:2px 8px;border-radius:4px;font-weight:bold;font-size:12px; }
    .badge-alta     { background:#FFD70022;color:#FFD700;padding:2px 8px;border-radius:4px;font-size:12px; }
    .badge-normal   { background:#4CAF5022;color:#81C784;padding:2px 8px;border-radius:4px;font-size:12px; }

    .device-code {
        background:#0D0F14;border:2px solid #00CFFF;border-radius:8px;
        padding:14px 20px;font-family:'Space Mono',monospace;
        font-size:26px;font-weight:700;color:#00CFFF;
        letter-spacing:0.15em;text-align:center;margin:12px 0;
    }
    .stButton > button[kind="primary"] {
        background:linear-gradient(135deg,#00CFFF,#0080FF);color:#0D0F14;
        font-weight:700;font-family:'Space Mono',monospace;
        border:none;border-radius:6px;padding:10px 24px;width:100%;
    }
    .stTabs [data-baseweb="tab-list"] { background:#13161D;border-radius:8px 8px 0 0;gap:4px; }
    .stTabs [data-baseweb="tab"]      { color:#7A8099;font-family:'Space Mono',monospace;font-size:12px; }
    .stTabs [aria-selected="true"]    { color:#00CFFF !important;border-bottom:2px solid #00CFFF; }

    .forecast-card {
        background:#13161D;border:1px solid #1E2330;border-radius:10px;
        padding:14px 10px;text-align:center;
    }
    .forecast-card.risco { border-color:#FF6B2D55; }

    iframe { color-scheme: light only !important; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────
# SESSION STATE
# ──────────────────────────────────────────────
_defaults = {
    "df_consolidado": None,
    "df_rota":        None,
    "df_ss":          None,
    "ss_map":         {},
    "ss_abertas_set": set(),
    "weather_map":    {},
    "resumo":         {},
    "fabric_authed":  False,
    "fabric_token":   None,
    "fabric_user":    None,
    "_device_flow":   None,
    "_msal_app":      None,
    "_msal_token_cache": None,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if not is_authenticated():
    tentar_login_silencioso()


# ──────────────────────────────────────────────
# SIDEBAR — LOGIN & FILTROS
# ──────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="app-title">⚡ INSPEÇÃO<br>LINHAS DE TRANSMISSÃO</div>', unsafe_allow_html=True)

    # ── Login ──
    if not is_authenticated():
        st.markdown("### 🔐 Login Energisa")

        if st.session_state["_device_flow"] is None:
            st.caption("Autentique com seu e-mail corporativo (suporta MFA).")
            if st.button("🔑 Iniciar Login", type="primary"):
                with st.spinner("Gerando código de acesso..."):
                    try:
                        flow = iniciar_device_flow()
                        st.session_state["_device_flow"] = flow
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")
        else:
            flow = st.session_state["_device_flow"]
            code = flow.get("user_code", "")
            url  = flow.get("verification_uri", "https://microsoft.com/devicelogin")

            st.markdown("#### Passo 1 — Copie o código")
            st.markdown(f'<div class="device-code">{code}</div>', unsafe_allow_html=True)
            st.markdown(f"#### Passo 2 — [Abra a página de login]({url})")
            st.markdown("#### Passo 3 — Confirme aqui")

            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ Já autentiquei", type="primary"):
                    with st.spinner("Verificando..."):
                        try:
                            concluir_login()
                            st.success("✅ Autenticado!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ {e}")
            with c2:
                if st.button("↩️ Reiniciar"):
                    st.session_state["_device_flow"] = None
                    st.rerun()

        st.stop()

    # ── Usuário logado ──
    st.success(f"✅ {st.session_state['fabric_user']}")
    if st.button("Sair"):
        logout()
        st.rerun()

    st.divider()
    st.markdown("### 🔍 Filtros")

    try:
        opcoes = get_filter_options(_sid=sid_atual())
        empresas = ["Todas"] + opcoes["empresas"]
        instalacoes_por_empresa = opcoes["instalacoes_por_empresa"]
    except Exception:
        empresas = ["Todas"]
        instalacoes_por_empresa = {}

    empresa_sel    = st.selectbox("Empresa",    empresas)
    _ins_lista     = (
        ["Todas"] + instalacoes_por_empresa.get(empresa_sel, [])
        if empresa_sel != "Todas"
        else ["Todas"] + sorted({i for lst in instalacoes_por_empresa.values() for i in lst})
    )
    instalacao_sel = st.selectbox("Instalação", _ins_lista)

    st.divider()
    st.markdown("### ⚙️ Parâmetros")

    max_os           = st.slider("Máximo de OS na rota", 5, 50, 20, step=5)
    forcar_atrasadas = st.toggle("🚨 Garantir ATRASADAS na rota", value=True)
    modo_conservador = st.toggle("🌦️ Destacar risco climático", value=True,
                                  help="Clima NÃO remove OS — apenas destaca visualmente")

    st.markdown("##### 🧠 Otimização inteligente")
    metodo_cluster = st.radio(
        "Método de clusterização",
        ["dbscan", "kmeans"],
        index=0,
        horizontal=True,
        help="DBSCAN: automático por densidade (recomendado). KMeans: k fixo.",
    )
    n_clusters_kmeans = 5
    if metodo_cluster == "kmeans":
        n_clusters_kmeans = st.slider("Número de clusters (k)", 2, 10, 5)
    usar_dois_opt = st.toggle("⚡ Melhoria 2-opt na rota", value=True,
                               help="Reduz ~5-15% a distância total. Levemente mais lento.")

    # Ponto de partida
    st.markdown("##### 📍 Ponto de partida (opcional)")
    _emp_key = None if empresa_sel    == "Todas" else empresa_sel
    _ins_key = None if instalacao_sel == "Todas" else instalacao_sel

    _filtro_atual = f"{_emp_key}|{_ins_key}"
    if st.session_state.get("_filtro_anterior") != _filtro_atual:
        st.session_state["_filtro_anterior"] = _filtro_atual
        st.session_state["_sel_partida"]     = 0

    df_torres_partida = pd.DataFrame()
    if _ins_key:
        try:
            df_torres_partida = load_torres_por_instalacao(_emp_key, _ins_key, _sid=sid_atual())
        except Exception:
            pass

    ponto_partida = None
    if df_torres_partida.empty:
        if _ins_key:
            st.caption("Nenhuma torre encontrada.")
        else:
            st.caption("Selecione uma instalação para escolher o ponto de partida.")
    else:
        opcoes_partida = ["— Início automático —"] + [
            f"Torre {r['NUM_TORRE']} — {r['COD_ATIVO']}"
            for _, r in df_torres_partida.iterrows()
        ]
        sel = st.selectbox("Torre de partida", opcoes_partida,
                           index=st.session_state.get("_sel_partida", 0))
        st.session_state["_sel_partida"] = opcoes_partida.index(sel)
        if sel != opcoes_partida[0]:
            idx   = opcoes_partida.index(sel) - 1
            rp    = df_torres_partida.iloc[idx]
            ponto_partida = (float(rp["LATITUDE"]), float(rp["LONGITUDE"]))

    st.divider()
    gerar = st.button("🚀 Gerar Rota Otimizada", type="primary")


# ──────────────────────────────────────────────
# LÓGICA PRINCIPAL
# ──────────────────────────────────────────────
if gerar:
    with st.spinner("🔄 Carregando OS do Fabric..."):
        try:
            df_raw = load_inspecoes_consolidadas(
                empresa    = None if empresa_sel    == "Todas" else empresa_sel,
                instalacao = None if instalacao_sel == "Todas" else instalacao_sel,
                _sid       = sid_atual(),
            )
        except Exception as e:
            st.error(f"❌ Erro ao conectar ao Fabric: {e}")
            st.stop()

    if df_raw.empty:
        st.warning("Nenhuma OS encontrada com os filtros selecionados.")
        st.stop()

    # 1. Pipeline completo: urgência + clusterização + score híbrido
    df_priorizado  = priorizar(df_raw)
    # Re-clusterizar com o método escolhido pelo usuário (priorizar usa dbscan por padrão)
    from utils.routing import clusterizar, calcular_score_hibrido
    df_priorizado = clusterizar(df_priorizado, metodo=metodo_cluster, n_clusters=n_clusters_kmeans)
    df_priorizado = calcular_score_hibrido(df_priorizado)
    df_priorizado = df_priorizado.sort_values(["PRIORIDADE", "SCORE"], ascending=[True, False]).reset_index(drop=True)

    # 2. Seleção de OS com modo cluster
    df_selecionado = selecionar_os(df_priorizado, max_os, forcar_atrasadas)

    # 2b. Carregar torres com SS N1/N2 em aberto e garantir que entram na rota
    df_torres_ss_abertas = pd.DataFrame()
    with st.spinner("📋 Buscando torres com SS N1/N2 em aberto..."):
        try:
            df_torres_ss_abertas = load_torres_com_ss_abertas(
                empresa    = None if empresa_sel    == "Todas" else empresa_sel,
                instalacao = None if instalacao_sel == "Todas" else instalacao_sel,
                _sid       = sid_atual(),
            )
        except Exception as e:
            st.warning(f"⚠️ Não foi possível carregar torres com SS abertas: {e}")

    # Marcar torres com SS N1/N2 no df_priorizado para destaque no mapa
    ss_abertas_set: set = set()
    if not df_torres_ss_abertas.empty and "COD_ATIVO" in df_torres_ss_abertas.columns:
        ss_abertas_set = set(df_torres_ss_abertas["COD_ATIVO"].dropna().unique())

    df_priorizado["TEM_SS_ABERTA"] = df_priorizado["COD_ATIVO"].isin(ss_abertas_set)

    # Incluir torres com SS N1/N2 que não estão na seleção atual
    if not df_torres_ss_abertas.empty:
        ativos_ja_selecionados = set(df_selecionado["COD_ATIVO"].dropna().unique())
        torres_faltantes_ss = df_torres_ss_abertas[
            ~df_torres_ss_abertas["COD_ATIVO"].isin(ativos_ja_selecionados)
        ]

        if not torres_faltantes_ss.empty:
            # Enriquecer as torres com SS com colunas mínimas para a rota
            # (STATUS_PRAZO=SS_ABERTA, PRIORIDADE=2 se N1, 3 se N2)
            rows_extras = []
            for _, tr in torres_faltantes_ss.iterrows():
                nivel_min = int(tr.get("NIVEL_MIN_SS", 2))
                prioridade_ss = 2 if nivel_min == 1 else 3
                row_extra = {
                    "COD_ATIVO"     : tr["COD_ATIVO"],
                    "NUM_TORRE"     : tr.get("NUM_TORRE", "–"),
                    "LATITUDE"      : float(tr["LATITUDE"]),
                    "LONGITUDE"     : float(tr["LONGITUDE"]),
                    "EMPRESA"       : tr.get("EMPRESA", "–"),
                    "INSTALACAO"    : tr.get("INSTALACAO", "–"),
                    "SIGLA_EMPRESA" : tr.get("EMPRESA", "–"),
                    "CRITICIDADE"   : int(tr.get("CRITICIDADE_MIN", 2)),
                    "STATUS_PRAZO"  : f"SS N{nivel_min} ABERTA",
                    "PRIORIDADE"    : prioridade_ss,
                    "SCORE"         : 70.0 if nivel_min == 1 else 50.0,
                    "SCORE_URG"     : 70.0 if nivel_min == 1 else 50.0,
                    "CLUSTER"       : 0,
                    "DIAS_ATRASO"   : 0,
                    "DESC_NUMERO_OS": f"SS-{tr.get('COD_ATIVO', '')}",
                    "TEM_SS_ABERTA" : True,
                }
                rows_extras.append(row_extra)

            if rows_extras:
                df_extras = pd.DataFrame(rows_extras)
                df_selecionado = pd.concat([df_selecionado, df_extras], ignore_index=True)
                df_selecionado["TEM_SS_ABERTA"] = df_selecionado["COD_ATIVO"].isin(ss_abertas_set)
                n_adicionadas = len(rows_extras)
                st.info(
                    f"🗼 **{n_adicionadas} torre(s) com SS N1/N2 em aberto** foram adicionadas "
                    f"automaticamente à rota e estão destacadas no mapa com anel laranja."
                )
    else:
        df_selecionado["TEM_SS_ABERTA"] = df_selecionado["COD_ATIVO"].isin(ss_abertas_set)

    # 3. Consulta climática das OS candidatas (apoio — não filtra)
    weather_map: dict = {}
    with st.spinner(f"🌦️ Consultando clima para {len(df_selecionado)} OS..."):
        prog = st.progress(0)
        for i, (_, row) in enumerate(df_selecionado.iterrows()):
            weather_map[row["COD_ATIVO"]] = get_weather(row["LATITUDE"], row["LONGITUDE"])
            prog.progress((i + 1) / len(df_selecionado))
            time.sleep(0.04)
        prog.empty()

    # 4. Otimização da rota (routing baseado em OS + coords via COD_ATIVO)
    df_rota = otimizar_rota(df_selecionado, ponto_partida, usar_dois_opt=usar_dois_opt)
    # Propagar flag TEM_SS_ABERTA para df_rota
    if "TEM_SS_ABERTA" not in df_rota.columns:
        df_rota["TEM_SS_ABERTA"] = df_rota["COD_ATIVO"].isin(ss_abertas_set)

    # 5. Carregar SS (contexto operacional — não afeta prioridade/rota)
    df_ss  = pd.DataFrame()
    ss_map: dict = {}
    ativos_rota = tuple(df_priorizado["COD_ATIVO"].dropna().unique().tolist())
    if ativos_rota:
        with st.spinner(f"📋 Carregando SS para {len(ativos_rota)} ativos..."):
            try:
                df_ss = load_ss_por_ativos(ativos_rota, _sid=sid_atual())
                if not df_ss.empty:
                    for ativo, grupo in df_ss.groupby("COD_ATIVO"):
                        ss_map[ativo] = grupo.to_dict(orient="records")
            except Exception as e:
                st.warning(f"⚠️ SS não carregadas: {e}")

    # Contar SS atrasadas (STATUS_PRAZO == 'ATRASADA' ou DIAS_EM_ABERTO > PRAZO_DIAS)
    ss_atrasadas = 0
    if not df_ss.empty:
        if "STATUS_PRAZO" in df_ss.columns:
            # STATUS_PRAZO pode ser "ATRASADO", "ATRASADA", "ATRASADO (Xd)" — usamos substring
            ss_atrasadas = int(df_ss["STATUS_PRAZO"].str.upper().str.contains("ATRASAD", na=False).sum())
        if ss_atrasadas == 0 and "DIAS_EM_ABERTO" in df_ss.columns and "STATUS_SS" in df_ss.columns:
            # Fallback: SS abertas (não concluídas) com dias em aberto > 0
            abertas = df_ss["STATUS_SS"].str.upper().str.contains("PENDENTE|ABERTA|ABERTO", na=False)
            ss_atrasadas = int((pd.to_numeric(df_ss.loc[abertas, "DIAS_EM_ABERTO"], errors="coerce").fillna(0) > 0).sum())

    _resumo = resumo_rota(df_rota)
    _resumo["ss_atrasadas"]     = ss_atrasadas
    _resumo["torres_ss_abertas"] = len(ss_abertas_set)

    st.session_state.df_consolidado  = df_priorizado
    st.session_state.df_rota         = df_rota
    st.session_state.df_ss           = df_ss
    st.session_state.ss_map          = ss_map
    st.session_state.weather_map     = weather_map
    st.session_state.resumo          = _resumo
    st.session_state.ss_abertas_set  = ss_abertas_set



# ──────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────
df_consolidado = st.session_state.df_consolidado
df_rota        = st.session_state.df_rota
df_ss          = st.session_state.df_ss
ss_map         = st.session_state.ss_map
ss_abertas_set = st.session_state.get("ss_abertas_set", set())
weather_map    = st.session_state.weather_map
resumo         = st.session_state.resumo

if resumo:
    c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
    c1.metric("📋 OS na rota",          resumo["total_os"])
    c2.metric("🔴 OS Atrasadas",         resumo["os_atrasadas"])
    c3.metric("⚠️ SS Atrasadas",         resumo.get("ss_atrasadas", "—"))
    c4.metric("🗼 Torres c/ SS Abertas", resumo.get("torres_ss_abertas", "—"))
    c5.metric("📏 Distância total",      f"{resumo['distancia_total']} km")
    c6.metric("📍 Dist. média/salto",    f"{resumo.get('distancia_media', '-')} km")
    c7.metric("🗂️ Clusters",             resumo.get("n_clusters", "-"))
    c8.metric("📊 Score médio",          f"{resumo['score_medio']}%")
    st.divider()

tab_mapa, tab_rota, tab_os, tab_ss, tab_ss_empresa, tab_clima = st.tabs([
    "🗺️  Mapa",
    "📋  Rota",
    "📂  OS Detalhadas",
    "⚠️  SS",
    "📊  SS por Empresa",
    "🌦️  Clima 5 dias",
])


# ── TAB MAPA ──
with tab_mapa:
    if df_consolidado is not None and not df_consolidado.empty:
        _hash = f"{len(df_rota)}_{df_rota['COD_ATIVO'].iloc[0] if df_rota is not None and not df_rota.empty else 'x'}"
        mapa  = build_map(df=df_consolidado, df_rota=df_rota, weather_map=weather_map, ss_map=ss_map, ss_abertas_set=ss_abertas_set)
        st_folium(mapa, use_container_width=True, height=580,
                  key=f"mapa_{_hash}", returned_objects=[])
    else:
        st.info("👈 Configure os filtros e clique em **Gerar Rota Otimizada**.")
        import folium
        m = folium.Map(location=[-15.8, -47.9], zoom_start=4, tiles=None)
        folium.TileLayer(
            "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
            attr="Esri", name="🛰️ Satélite",
        ).add_to(m)
        st_folium(m, use_container_width=True, height=500, key="mapa_vazio", returned_objects=[])


# ── TAB ROTA ──
with tab_rota:
    if df_rota is not None and not df_rota.empty:
        st.markdown(f"#### Rota com **{len(df_rota)} OS** — {resumo['distancia_total']} km")

        # Cores por prioridade
        def _cor_prioridade_css(v):
            try:
                p = int(v)
                if p == int(Prioridade.MAXIMA): return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                if p == int(Prioridade.ALTA):   return "background:#FFD70022;color:#FFD700"
            except Exception: pass
            return ""

        def _label_prioridade(v):
            try:
                p = int(v)
                return {
                    int(Prioridade.MAXIMA): "🔴 ATRASADA",
                    int(Prioridade.ALTA):   "🟡 VENCE EM BREVE",
                    int(Prioridade.NORMAL): "🟢 NO PRAZO",
                }.get(p, str(v))
            except Exception:
                return str(v)

        _cols_rota = ["ORDEM_VISITA", "DESC_NUMERO_OS", "COD_ATIVO", "NUM_TORRE",
                       "SIGLA_EMPRESA", "INSTALACAO", "DESC_ESTADO",
                       "STATUS_PRAZO", "DATA_LIMITE", "DIAS_ATRASO",
                       "PRIORIDADE", "SCORE", "CLUSTER", "DIST_PROX_KM", "DIST_ACUM_KM"]
        df_exibir = df_rota[[c for c in _cols_rota if c in df_rota.columns]].copy()

        df_exibir["PRIORIDADE"]   = df_exibir["PRIORIDADE"].apply(_label_prioridade)
        df_exibir["DATA_LIMITE"]  = pd.to_datetime(df_exibir["DATA_LIMITE"], errors="coerce").dt.strftime("%d/%m/%Y")
        df_exibir["SCORE"]        = df_exibir["SCORE"].apply(lambda v: f"{float(v):.1f}%" if pd.notna(v) and v != "–" else "–")
        df_exibir["DIST_PROX_KM"] = df_exibir["DIST_PROX_KM"].apply(lambda v: f"{v:.1f}" if v else "–")
        df_exibir["DIST_ACUM_KM"] = df_exibir["DIST_ACUM_KM"].apply(lambda v: f"{v:.1f}" if v else "–")

        rename = {
            "ORDEM_VISITA":  "Ordem",    "DESC_NUMERO_OS": "OS",
            "COD_ATIVO":     "Ativo",    "NUM_TORRE":       "Torre",
            "SIGLA_EMPRESA": "Empresa",  "INSTALACAO":      "Instalação",
            "DESC_ESTADO":   "Estado",   "STATUS_PRAZO":    "Status",
            "DATA_LIMITE":   "Limite",   "DIAS_ATRASO":     "Atraso (d)",
            "PRIORIDADE":    "Prioridade","SCORE":           "Score",
            "CLUSTER":       "Cluster",  "DIST_PROX_KM":    "Dist→(km)",
            "DIST_ACUM_KM":  "Acum.(km)",
        }
        df_exibir = df_exibir.rename(columns=rename)

        styled = df_exibir.style.map(
            lambda v: "background:#FF2D2D22;color:#FF6B6B;font-weight:bold" if "ATRASADA" in str(v)
            else ("background:#FFD70022;color:#FFD700" if "VENCE" in str(v) else ""),
            subset=["Prioridade"]
        ).map(
            lambda v: "color:#FF6B6B;font-weight:bold" if int(v) > 0 else ""
            if pd.notna(v) and str(v).lstrip('-').isdigit() else "",
            subset=["Atraso (d)"]
        )

        st.dataframe(styled, use_container_width=True, hide_index=True)

        # Export Excel
        xlsx = _gerar_excel_rota(df_exibir, resumo)
        st.download_button("📥 Exportar Excel", xlsx, "rota_inspecao.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Gere a rota para ver o detalhamento.")


# ── TAB OS DETALHADAS ──
with tab_os:
    if df_consolidado is not None and not df_consolidado.empty:
        col_titulo, col_ajuda = st.columns([5, 1])
        with col_titulo:
            st.markdown(f"#### {len(df_consolidado)} OS consolidadas (todas, ordenadas por prioridade)")
        with col_ajuda:
            with st.popover("ℹ️ Como funciona?", use_container_width=True):
                st.markdown("""
### 🏷️ Prioridade

A prioridade é calculada automaticamente com base no prazo de cada OS:

| Valor | Nível | Critério |
|-------|-------|----------|
| **1** | 🔴 **ATRASADA** | `STATUS_PRAZO = 'ATRASADA'` ou `DATA_LIMITE` já vencida |
| **2** | 🟡 **ALTA** | `DATA_LIMITE` vence em **≤ 7 dias** |
| **3** | 🟢 **NORMAL** | Demais OS dentro do prazo |

---

### 📊 Score Híbrido (0 – 100)

O Score combina **urgência + localização geográfica** em três componentes:

```
Score = 55% × Score_Urgência
      + 30% × Bônus_Cluster
      + 15% × Bônus_Densidade
```

#### Score de Urgência (base)
```
base        = (4 − Prioridade) × 30       → máx 90 pts
bônus_atraso = min(atraso, 60) × 0,15     → máx  9 pts
bônus_vencer = (7 − dias_p_vencer)/7 × 10 → máx 10 pts  (só P=2)
```

| Prioridade | Base | + Atraso | + Vencimento |
|---|---|---|---|
| 🔴 ATRASADA (P=1) | 90 pts | até +9 | — |
| 🟡 ALTA (P=2) | 60 pts | até +9 | até +10 |
| 🟢 NORMAL (P=3) | 30 pts | até +9 | — |

#### Bônus Cluster
Proporção do cluster em relação ao maior cluster identificado (0–100). OS em clusters maiores recebem bônus mais alto — priorizando regiões com mais inspeções concentradas.

#### Bônus Densidade
Percentual de torres vizinhas num raio de **50 km**, normalizado (0–100). Incentiva rotas mais compactas geograficamente.

---

> **Exemplo:** OS atrasada há 20 dias, no maior cluster, região densa
> - Urgência: 90 + min(20,60)×0,15 = **93**
> - Score = 0,55×93 + 0,30×100 + 0,15×100 = **51 + 30 + 15 = 96**

---

> 💡 A tabela ordena por: **Prioridade ↑ → Score ↓**
""")

        col_filtro, col_vazio = st.columns([1, 3])
        with col_filtro:
            filtro_status = st.selectbox("Filtrar por status",
                                         ["Todos", "🔴 Atrasadas", "🟡 Alta", "🟢 Normal"])

        df_os = df_consolidado.copy()
        if filtro_status == "🔴 Atrasadas":
            df_os = df_os[df_os["PRIORIDADE"] == int(Prioridade.MAXIMA)]
        elif filtro_status == "🟡 Alta":
            df_os = df_os[df_os["PRIORIDADE"] == int(Prioridade.ALTA)]
        elif filtro_status == "🟢 Normal":
            df_os = df_os[df_os["PRIORIDADE"] == int(Prioridade.NORMAL)]

        colunas = ["DESC_NUMERO_OS", "COD_ATIVO", "NUM_TORRE", "SIGLA_EMPRESA",
                   "INSTALACAO", "STATUS_PRAZO", "DATA_LIMITE", "DIAS_ATRASO",
                   "PRIORIDADE", "SCORE", "DESC_ESTADO"]
        df_os_exib = df_os[[c for c in colunas if c in df_os.columns]].copy()
        df_os_exib["DATA_LIMITE"] = pd.to_datetime(df_os_exib["DATA_LIMITE"], errors="coerce").dt.strftime("%d/%m/%Y")
        if "SCORE" in df_os_exib.columns:
            df_os_exib["SCORE"] = df_os_exib["SCORE"].apply(
                lambda v: f"{float(v):.1f}%" if pd.notna(v) else "–"
            )

        st.dataframe(df_os_exib, use_container_width=True, hide_index=True)

        xlsx_os = _gerar_excel_os(df_os_exib)
        st.download_button("📥 Exportar OS Excel", xlsx_os, "os_detalhadas.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_os")
    else:
        st.info("Gere a rota para ver as OS detalhadas.")



# ── TAB SS ──
with tab_ss:
    if df_ss is not None and not df_ss.empty:
        n_ativos_ss = df_ss["COD_ATIVO"].nunique() if "COD_ATIVO" in df_ss.columns else 0
        n1 = int((df_ss["NIVEL_SS"] == 1).sum()) if "NIVEL_SS" in df_ss.columns else 0
        n2 = int((df_ss["NIVEL_SS"] == 2).sum()) if "NIVEL_SS" in df_ss.columns else 0

        col_t, col_aj = st.columns([5, 1])
        with col_t:
            st.markdown(f"#### ⚠️ {len(df_ss)} Solicitações de Serviço — {n_ativos_ss} ativos")
            st.caption(
                f"🔴 **Nível 1:** {n1}  &nbsp;&nbsp; 🟡 **Nível 2:** {n2}  &nbsp;&nbsp; "
                "SS são contexto operacional — **não afetam prioridade nem rota**."
            )
        with col_aj:
            with st.popover("ℹ️ O que são SS?", use_container_width=True):
                st.markdown("""
### ⚠️ Solicitações de Serviço (SS)

As SS registram **defeitos existentes** em torres de transmissão.

#### Como são usadas neste sistema
| Uso | Comportamento |
|-----|--------------|
| Priorização de OS | ❌ Não afeta |
| Inclusão/exclusão da rota | ❌ Não afeta |
| Contexto na inspeção | ✅ Exibidas junto à OS |
| Popup no mapa | ✅ Visíveis ao clicar na torre |

#### Níveis exibidos
| Nível | Criticidade |
|-------|-------------|
| **1** | 🔴 Alta — defeito crítico |
| **2** | 🟡 Média — defeito relevante |

> O inspetor vê as SS associadas à torre ao planejar a visita,
> podendo se preparar com ferramentas e peças adequadas.
""")

        # Filtros
        col_f1, col_f2, col_f3 = st.columns([1, 1, 2])
        with col_f1:
            nivel_filtro = st.selectbox(
                "Nível SS", ["Todos", "🔴 Nível 1", "🟡 Nível 2"], key="ss_nivel"
            )
        with col_f2:
            status_opts = ["Todos"]
            if "STATUS_SS" in df_ss.columns:
                status_opts += sorted(df_ss["STATUS_SS"].dropna().unique().tolist())
            status_filtro = st.selectbox("Status", status_opts, key="ss_status")
        with col_f3:
            ativo_opts = ["Todos"] + sorted(df_ss["COD_ATIVO"].dropna().unique().tolist())
            ativo_filtro = st.selectbox("Ativo (COD_ATIVO)", ativo_opts, key="ss_ativo")

        df_ss_exib = df_ss.copy()
        _col_nivel = "NIVEL_CRITICIDADE" if "NIVEL_CRITICIDADE" in df_ss_exib.columns else "NIVEL_SS"
        if nivel_filtro == "🔴 Nível 1":
            df_ss_exib = df_ss_exib[pd.to_numeric(df_ss_exib.get(_col_nivel), errors="coerce") == 1]
        elif nivel_filtro == "🟡 Nível 2":
            df_ss_exib = df_ss_exib[pd.to_numeric(df_ss_exib.get(_col_nivel), errors="coerce") == 2]
        if status_filtro != "Todos" and "STATUS_SS" in df_ss_exib.columns:
            df_ss_exib = df_ss_exib[df_ss_exib["STATUS_SS"] == status_filtro]
        if ativo_filtro != "Todos":
            df_ss_exib = df_ss_exib[df_ss_exib["COD_ATIVO"] == ativo_filtro]

        # Formatar data
        if "DATA_ABERTURA" in df_ss_exib.columns:
            df_ss_exib = df_ss_exib.copy()
            df_ss_exib["DATA_ABERTURA"] = pd.to_datetime(
                df_ss_exib["DATA_ABERTURA"], errors="coerce"
            ).dt.strftime("%d/%m/%Y")

        # Colorir por nível — suporta NIVEL_CRITICIDADE (real) ou NIVEL_SS (legado)
        _col_nivel_tab = "NIVEL_CRITICIDADE" if "NIVEL_CRITICIDADE" in df_ss_exib.columns else "NIVEL_SS"
        _colunas_ss = [c for c in [
            "COD_SS", "COD_ATIVO", _col_nivel_tab, "TIPO_DEFEITO",
            "DESC_SS", "STATUS_SS", "STATUS_PRAZO", "DATA_ABERTURA"
        ] if c in df_ss_exib.columns]

        def _cor_nivel_ss(val):
            try:
                n = int(val)
                if n == 1: return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                if n == 2: return "background:#FFD70022;color:#FFD700"
            except Exception:
                pass
            return ""

        def _cor_status_prazo(val):
            v = str(val).upper()
            if "ATRASAD" in v: return "color:#FF6B6B;font-weight:bold"
            return ""

        styled_ss = df_ss_exib[_colunas_ss].style
        if _col_nivel_tab in _colunas_ss:
            styled_ss = styled_ss.map(_cor_nivel_ss, subset=[_col_nivel_tab])
        if "STATUS_PRAZO" in _colunas_ss:
            styled_ss = styled_ss.map(_cor_status_prazo, subset=["STATUS_PRAZO"])

        st.dataframe(styled_ss, use_container_width=True, hide_index=True)

        # Cruzamento com a rota: quais torres da rota têm SS?
        # Usa df_ss diretamente (não ss_map) para garantir que todas as torres apareçam
        if df_rota is not None and not df_rota.empty and "COD_ATIVO" in df_ss.columns:
            # Contagem de SS por ativo a partir do df_ss completo
            ss_por_ativo = df_ss.groupby("COD_ATIVO").size().to_dict()

            ativos_rota_set = set(df_rota["COD_ATIVO"].dropna().unique())
            ativos_ss_set   = set(df_ss["COD_ATIVO"].dropna().unique())
            ativos_cruzados = ativos_rota_set & ativos_ss_set

            if ativos_cruzados:
                st.divider()
                st.markdown(f"#### 🔗 Torres da rota com SS vinculadas ({len(ativos_cruzados)})")
                df_cruzado = df_rota[df_rota["COD_ATIVO"].isin(ativos_cruzados)][
                    [c for c in ["ORDEM_VISITA", "COD_ATIVO", "NUM_TORRE", "INSTALACAO",
                                 "STATUS_PRAZO", "PRIORIDADE"] if c in df_rota.columns]
                ].copy()
                df_cruzado["Qtd SS"] = df_cruzado["COD_ATIVO"].map(
                    lambda a: f"{ss_por_ativo.get(a, 0)} SS"
                )
                # Verificar SS atrasadas por ativo
                if "STATUS_PRAZO" in df_ss.columns:
                    ss_atr_por_ativo = df_ss[
                        df_ss["STATUS_PRAZO"].str.upper().str.contains("ATRASAD", na=False)
                    ].groupby("COD_ATIVO").size().to_dict()
                    df_cruzado["SS Atrasadas"] = df_cruzado["COD_ATIVO"].map(
                        lambda a: ss_atr_por_ativo.get(a, 0)
                    )
                st.dataframe(df_cruzado, use_container_width=True, hide_index=True)
                st.caption("ℹ️ Essas torres têm defeitos registrados. O inspetor deve verificá-los durante a visita.")

        # Export
        buf_ss = io.BytesIO()
        df_ss_exib[_colunas_ss].to_excel(buf_ss, index=False, engine="openpyxl")
        st.download_button(
            "📥 Exportar SS Excel", buf_ss.getvalue(), "ss_vinculadas.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ss",
        )

    elif df_consolidado is not None:
        st.info("Nenhuma SS de nível 1 ou 2 encontrada para os ativos carregados.")
    else:
        st.info("Gere a rota para ver as SS vinculadas.")



# ── TAB SS POR EMPRESA ──
with tab_ss_empresa:
    st.markdown("#### 📊 SS por Empresa e Instalação — Painel de Apoio")
    st.caption(
        "Painel independente dos filtros do sidebar. "
        "Use-o para identificar quais empresas e instalações concentram mais SS abertas (N1–N2) e atrasadas, "
        "e então aplique esses filtros na sidebar para gerar a rota priorizada."
    )

    # ── Filtros próprios desta aba ──────────────────────────────────────
    with st.expander("🔍 Filtros", expanded=True):
        col_fe, col_fi, col_fn, col_fb = st.columns([1, 1, 1, 1])

        with col_fe:
            # Opções de empresa vindas do cache de filtros (mesma fonte do sidebar, mas independente)
            try:
                _opcoes_emp = get_filter_options(_sid=sid_atual())
                _empresas_ss = ["Todas"] + _opcoes_emp["empresas"]
                _inst_por_emp = _opcoes_emp["instalacoes_por_empresa"]
            except Exception:
                _empresas_ss = ["Todas"]
                _inst_por_emp = {}
            ss_emp_sel = st.selectbox("Empresa", _empresas_ss, key="sse_empresa")

        with col_fi:
            if ss_emp_sel != "Todas":
                _inst_opts = ["Todas"] + _inst_por_emp.get(ss_emp_sel, [])
            else:
                _inst_opts = ["Todas"] + sorted({i for lst in _inst_por_emp.values() for i in lst})
            ss_ins_sel = st.selectbox("Instalação", _inst_opts, key="sse_instalacao")

        with col_fn:
            ss_nivel_max = st.selectbox(
                "Criticidade máxima",
                [1, 2, 3, 4, 5, 6],
                index=1,
                format_func=lambda n: f"Nível 1 a {n}",
                key="sse_nivel_max",
            )

        with col_fb:
            st.markdown("<br>", unsafe_allow_html=True)
            buscar_ss_emp = st.button("🔎 Buscar SS", type="primary", key="sse_buscar")

    # ── Inicializa estado local ──────────────────────────────────────────
    if "df_ss_empresa" not in st.session_state:
        st.session_state["df_ss_empresa"] = None
    if buscar_ss_emp:
        with st.spinner("🔄 Carregando SS + dados de torres..."):
            try:
                st.session_state["df_ss_empresa"] = load_ss_por_empresa(
                    empresa    = None if ss_emp_sel == "Todas" else ss_emp_sel,
                    instalacao = None if ss_ins_sel == "Todas" else ss_ins_sel,
                    nivel_max  = ss_nivel_max,
                    _sid       = sid_atual(),
                )
            except Exception as e:
                st.error(f"❌ Erro ao carregar SS: {e}")
                st.session_state["df_ss_empresa"] = None

    df_sse = st.session_state.get("df_ss_empresa")

    if df_sse is not None and not df_sse.empty:

        # ── Métricas rápidas ─────────────────────────────────────────────
        _n_empresas = df_sse["EMPRESA"].nunique() if "EMPRESA" in df_sse.columns else "—"
        _n_inst     = df_sse["INSTALACAO"].nunique() if "INSTALACAO" in df_sse.columns else "—"
        _n_ss       = df_sse["COD_SS"].nunique() if "COD_SS" in df_sse.columns else len(df_sse)
        _n_torres   = df_sse["COD_ATIVO"].nunique() if "COD_ATIVO" in df_sse.columns else "—"
        _atr_mask   = df_sse["STATUS_PRAZO"].str.upper().str.startswith("ATRASADO", na=False) if "STATUS_PRAZO" in df_sse.columns else pd.Series(False, index=df_sse.index)
        _n_atr      = int(_atr_mask.sum())
        _n1_count   = int((df_sse["NIVEL_CRITICIDADE"] == 1).sum()) if "NIVEL_CRITICIDADE" in df_sse.columns else "—"

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("🏢 Empresas",      _n_empresas)
        m2.metric("📍 Instalações",   _n_inst)
        m3.metric("🗼 Torres",        _n_torres)
        m4.metric("📋 SS abertas",    _n_ss)
        m5.metric("🔴 SS Nível 1",   _n1_count)
        m6.metric("⏰ SS Atrasadas",  _n_atr)

        st.divider()

        # ── Escolha de visualização ──────────────────────────────────────
        vis_col, ord_col, _ = st.columns([1, 1, 2])
        with vis_col:
            vis_mode = st.radio(
                "Visualização",
                ["📊 Resumo por instalação", "📋 SS individuais"],
                horizontal=True,
                key="sse_vis",
            )
        with ord_col:
            ord_mode = st.selectbox(
                "Ordenar por",
                ["SS Atrasadas ↓", "SS Nível 1 ↓", "Total SS ↓", "Pior saldo (dias) ↑"],
                key="sse_ord",
            )

        # ── MODO RESUMO ──────────────────────────────────────────────────
        if vis_mode == "📊 Resumo por instalação":
            grp = ["EMPRESA", "INSTALACAO"]
            grp = [c for c in grp if c in df_sse.columns]

            agg = df_sse.groupby(grp).agg(
                Torres      = ("COD_ATIVO",          "nunique"),
                Total_SS    = ("COD_SS",              "nunique"),
                N1          = ("NIVEL_CRITICIDADE",   lambda x: (x == 1).sum()),
                N2          = ("NIVEL_CRITICIDADE",   lambda x: (x == 2).sum()),
                Atrasadas   = ("STATUS_PRAZO",        lambda x: x.str.upper().str.startswith("ATRASADO", na=False).sum()),
                Pior_Saldo  = ("SALDO_DIAS",          "min"),
                Max_Aberto  = ("DIAS_EM_ABERTO",      "max"),
            ).reset_index()

            # Renomear
            agg = agg.rename(columns={
                "EMPRESA": "Empresa", "INSTALACAO": "Instalação",
                "Torres": "Torres", "Total_SS": "Total SS",
                "N1": "🔴 N1", "N2": "🟡 N2",
                "Atrasadas": "⏰ Atrasadas",
                "Pior_Saldo": "Pior Saldo (d)",
                "Max_Aberto": "Máx. Dias Aberto",
            })

            # Ordenação
            sort_map = {
                "SS Atrasadas ↓":       ("⏰ Atrasadas",     False),
                "SS Nível 1 ↓":        ("🔴 N1",            False),
                "Total SS ↓":           ("Total SS",          False),
                "Pior saldo (dias) ↑":  ("Pior Saldo (d)",   True),
            }
            sort_col, sort_asc = sort_map.get(ord_mode, ("⏰ Atrasadas", False))
            agg = agg.sort_values(sort_col, ascending=sort_asc).reset_index(drop=True)

            def _style_resumo(val):
                try:
                    v = float(val)
                    if v > 0:
                        return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                except Exception:
                    pass
                return ""

            def _style_saldo(val):
                try:
                    v = float(val)
                    if v < 0:
                        return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                    if v <= 7:
                        return "background:#FFD70022;color:#FFD700"
                except Exception:
                    pass
                return ""

            styled_agg = agg.style \
                .map(_style_resumo, subset=["⏰ Atrasadas", "🔴 N1"]) \
                .map(_style_saldo,  subset=["Pior Saldo (d)"])

            st.dataframe(styled_agg, use_container_width=True, hide_index=True)

            st.info(
                "💡 **Como usar:** Identifique as instalações mais críticas acima, "
                "depois vá ao **sidebar → Filtros** e selecione a Empresa e Instalação "
                "correspondente para gerar a rota priorizada."
            )

        # ── MODO SS INDIVIDUAIS ──────────────────────────────────────────
        else:
            # Filtros adicionais rápidos dentro da aba
            fc1, fc2, fc3 = st.columns([1, 1, 2])
            with fc1:
                nv_opts = ["Todos"] + sorted(df_sse["NIVEL_CRITICIDADE"].dropna().unique().astype(int).tolist())
                nv_f = st.selectbox("Nível", nv_opts, key="sse_ind_nivel",
                                    format_func=lambda v: f"Nível {v}" if v != "Todos" else "Todos")
            with fc2:
                st_opts = ["Todos", "Atrasadas", "No prazo", "Sem SLA"]
                st_f = st.selectbox("Status prazo", st_opts, key="sse_ind_status")

            df_ind = df_sse.copy()
            if nv_f != "Todos":
                df_ind = df_ind[df_ind["NIVEL_CRITICIDADE"] == int(nv_f)]
            if st_f == "Atrasadas":
                df_ind = df_ind[df_ind["STATUS_PRAZO"].str.upper().str.startswith("ATRASADO", na=False)]
            elif st_f == "No prazo":
                df_ind = df_ind[df_ind["STATUS_PRAZO"].str.upper().str.startswith("FALTAM", na=False)]
            elif st_f == "Sem SLA":
                df_ind = df_ind[df_ind["STATUS_PRAZO"].str.upper() == "SEM SLA"]

            # Ordenação
            sort_map_ind = {
                "SS Atrasadas ↓":      ("SALDO_DIAS",          True),
                "SS Nível 1 ↓":       ("NIVEL_CRITICIDADE",   True),
                "Total SS ↓":          ("DIAS_EM_ABERTO",      False),
                "Pior saldo (dias) ↑": ("SALDO_DIAS",          True),
            }
            sc, sa = sort_map_ind.get(ord_mode, ("SALDO_DIAS", True))
            if sc in df_ind.columns:
                df_ind = df_ind.sort_values(sc, ascending=sa)

            # Colunas para exibição
            _cols_ind = [c for c in [
                "EMPRESA", "INSTALACAO", "NUM_TORRE", "COD_ATIVO",
                "COD_SS", "NIVEL_CRITICIDADE", "codigo_do_defeito",
                "descricao_do_defeito", "STATUS_PRAZO", "SALDO_DIAS",
                "DIAS_EM_ABERTO", "DATA_REQUISICAO", "DATA_LIMITE",
                "ESTADO_SS", "NOME_PRIORIDADE", "TEXT_OBSERVACAO",
            ] if c in df_ind.columns]

            # Formatar datas
            for dc in ["DATA_REQUISICAO", "DATA_LIMITE"]:
                if dc in df_ind.columns:
                    df_ind[dc] = pd.to_datetime(df_ind[dc], errors="coerce").dt.strftime("%d/%m/%Y")

            def _cor_ind(val):
                try:
                    n = int(val)
                    if n == 1: return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                    if n == 2: return "background:#FFD70022;color:#FFD700"
                except Exception:
                    s = str(val).upper()
                    if s.startswith("ATRASADO"): return "background:#FF2D2D22;color:#FF6B6B;font-weight:bold"
                    if s.startswith("FALTAM"):
                        try:
                            dias = int(s.split()[1])
                            if dias <= 7: return "background:#FFD70022;color:#FFD700"
                        except Exception:
                            pass
                return ""

            styled_ind = df_ind[_cols_ind].style
            if "NIVEL_CRITICIDADE" in _cols_ind:
                styled_ind = styled_ind.map(_cor_ind, subset=["NIVEL_CRITICIDADE"])
            if "STATUS_PRAZO" in _cols_ind:
                styled_ind = styled_ind.map(_cor_ind, subset=["STATUS_PRAZO"])

            st.dataframe(styled_ind, use_container_width=True, hide_index=True)

        # ── Export ───────────────────────────────────────────────────────
        st.divider()
        buf_sse = io.BytesIO()
        df_sse.to_excel(buf_sse, index=False, engine="openpyxl")
        st.download_button(
            "📥 Exportar Excel", buf_sse.getvalue(),
            "ss_por_empresa.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_ss_empresa",
        )

    elif df_sse is not None and df_sse.empty:
        st.info("Nenhuma SS encontrada com os critérios selecionados.")
    else:
        st.info("👆 Selecione os filtros acima e clique em **Buscar SS** para carregar o painel.")


# ── TAB CLIMA 5 DIAS ──
with tab_clima:
    if df_rota is not None and not df_rota.empty:
        st.markdown("#### 🌦️ Previsão de 5 dias — OS da rota")
        st.caption("⚠️ Clima é apenas apoio — não remove OS da rota.")

        # Seleciona OS para ver previsão
        os_opcoes = df_rota["COD_ATIVO"].tolist()
        os_sel    = st.selectbox("Selecione o ativo:", os_opcoes,
                                  format_func=lambda c: f"{c} — Torre {df_rota.loc[df_rota['COD_ATIVO']==c,'NUM_TORRE'].values[0] if len(df_rota.loc[df_rota['COD_ATIVO']==c])>0 else '?'}")

        row_sel = df_rota[df_rota["COD_ATIVO"] == os_sel].iloc[0]
        lat, lon = float(row_sel["LATITUDE"]), float(row_sel["LONGITUDE"])

        with st.spinner("Buscando previsão..."):
            forecast = get_forecast_5d(lat, lon)
            clima_atual = weather_map.get(os_sel) or get_weather(lat, lon)

        # Clima atual
        if clima_atual.get("ok"):
            badge = "⛔ RISCO OPERACIONAL" if clima_atual["risco"] else "✅ Condições OK"
            st.markdown(f"""
            **Clima atual** &nbsp; {badge}
            | 🌡️ {clima_atual['temperatura']}°C | 💧 {clima_atual['umidade']}% | 💨 {clima_atual['vento_kmh']} km/h | 🌧️ {clima_atual['chuva_mm']} mm/h
            """)

        st.divider()

        # Cards de 5 dias
        if forecast:
            cols = st.columns(min(len(forecast), 5))
            for i, dia in enumerate(forecast[:5]):
                with cols[i]:
                    card_class = "forecast-card risco" if dia["risco"] else "forecast-card"
                    risco_badge = "⛔" if dia["risco"] else "✅"
                    st.markdown(f"""
                    <div class="{card_class}">
                        <div style='font-size:16px;font-weight:bold;color:#00CFFF'>{dia['data']}</div>
                        <div style='font-size:22px;margin:6px 0'>
                            <img src='https://openweathermap.org/img/wn/{dia['icone']}@2x.png' width='48'>
                        </div>
                        <div style='font-size:13px;color:#E8EAF0;margin-bottom:4px'>{dia['descricao']}</div>
                        <div style='font-size:12px;color:#7A8099'>
                            🌡️ {dia['temp_min']}° / {dia['temp_max']}°C<br>
                            💨 {dia['vento_kmh']} km/h<br>
                            🌧️ {dia['chuva_mm']} mm<br>
                            {risco_badge}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.warning("Previsão não disponível para esta localização.")

        st.divider()

        # Resumo climático de todas as OS da rota
        st.markdown("#### Resumo climático — todas as OS da rota")
        dados_clima = []
        for cod, info in weather_map.items():
            if not isinstance(info, dict):
                continue
            row_match = df_rota[df_rota["COD_ATIVO"] == cod]
            torre = row_match["NUM_TORRE"].values[0] if not row_match.empty else "–"
            dados_clima.append({
                "Ativo":       cod,
                "Torre":       torre,
                "Condição":    info.get("descricao", "N/D") if info.get("ok") else f"Erro: {info.get('erro','')}",
                "Temp (°C)":   round(info["temperatura"], 1) if info.get("ok") else "–",
                "Umidade (%)": info.get("umidade", "–") if info.get("ok") else "–",
                "Vento (km/h)":info.get("vento_kmh", "–") if info.get("ok") else "–",
                "Chuva (mm/h)":info.get("chuva_mm", "–") if info.get("ok") else "–",
                "Status":      "⛔ RISCO" if info.get("risco") else "✅ OK",
            })

        if dados_clima:
            df_clima_tab = pd.DataFrame(dados_clima)
            st.dataframe(
                df_clima_tab.style.map(
                    lambda v: "color:#FF6B6B;font-weight:bold" if "RISCO" in str(v) else "",
                    subset=["Status"]
                ),
                use_container_width=True, hide_index=True,
            )
    else:
        st.info("Gere a rota para ver a previsão climática.")
